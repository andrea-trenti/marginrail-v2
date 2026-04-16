from __future__ import annotations

import io
import json
import subprocess
import sys
import types
from importlib.util import module_from_spec, spec_from_file_location
from pathlib import Path
from zipfile import ZipFile

import pandas as pd
import pytest
from openpyxl import Workbook


class _DummyContext:
    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False


class _FakeStreamlit(types.ModuleType):
    def __init__(self):
        super().__init__("streamlit")
        self.session_state = {}

    def set_page_config(self, *args, **kwargs):
        return None

    def title(self, *args, **kwargs):
        return None

    def caption(self, *args, **kwargs):
        return None

    def subheader(self, *args, **kwargs):
        return None

    def success(self, *args, **kwargs):
        return None

    def error(self, *args, **kwargs):
        return None

    def write(self, *args, **kwargs):
        return None

    def info(self, *args, **kwargs):
        return None

    def dataframe(self, *args, **kwargs):
        return None

    def metric(self, *args, **kwargs):
        return None

    def code(self, *args, **kwargs):
        return None

    def file_uploader(self, *args, **kwargs):
        return None

    def button(self, *args, **kwargs):
        return False

    def download_button(self, *args, **kwargs):
        return False

    def multiselect(self, *args, **kwargs):
        default = kwargs.get("default")
        return [] if default is None else default

    def columns(self, spec, *args, **kwargs):
        if isinstance(spec, int):
            count = spec
        else:
            count = len(spec)
        return [_DummyContext() for _ in range(count)]

    def spinner(self, *args, **kwargs):
        return _DummyContext()

    def expander(self, *args, **kwargs):
        return _DummyContext()


@pytest.fixture(scope="session")
def marginrail_module():
    project_root = Path(__file__).resolve().parents[1]
    app_path = project_root / "app.py"

    fake_streamlit = _FakeStreamlit()
    previous_streamlit = sys.modules.get("streamlit")
    sys.modules["streamlit"] = fake_streamlit

    spec = spec_from_file_location("marginrail_app", app_path)
    module = module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)

    if previous_streamlit is not None:
        sys.modules["streamlit"] = previous_streamlit
    else:
        del sys.modules["streamlit"]

    return module


@pytest.fixture()
def valid_excel_bytes(marginrail_module):
    return build_workbook_bytes(sorted(marginrail_module.REQUIRED_SHEETS))


def build_workbook_bytes(sheet_names: list[str]) -> bytes:
    wb = Workbook()
    first = wb.active
    first.title = sheet_names[0]
    first["A1"] = "placeholder"
    for name in sheet_names[1:]:
        ws = wb.create_sheet(title=name)
        ws["A1"] = "placeholder"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_report_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "ok"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_validate_excel_structure_rejects_invalid_bytes(marginrail_module):
    result = marginrail_module.validate_excel_structure(b"not-an-xlsx")

    assert result["ok"] is False
    assert "non è un Excel .xlsx valido" in result["message"]
    assert result["missing"] == sorted(marginrail_module.REQUIRED_SHEETS)
    assert result["found"] == []


def test_validate_excel_structure_detects_missing_required_sheet(marginrail_module):
    missing_sheet = "Promo_2025"
    available = sorted(marginrail_module.REQUIRED_SHEETS - {missing_sheet})
    workbook_bytes = build_workbook_bytes(available)

    result = marginrail_module.validate_excel_structure(workbook_bytes)

    assert result["ok"] is False
    assert result["message"] == "Mancano uno o più fogli obbligatori."
    assert result["missing"] == [missing_sheet]
    assert set(result["found"]) == set(available)


def test_validate_excel_structure_accepts_complete_workbook(marginrail_module, valid_excel_bytes):
    result = marginrail_module.validate_excel_structure(valid_excel_bytes)

    assert result["ok"] is True
    assert result["missing"] == []
    assert set(result["found"]) == set(marginrail_module.REQUIRED_SHEETS)


@pytest.mark.parametrize(
    ("stderr", "stdout", "expected"),
    [
        (
            "Sheet 'Vendite_2025' non trovata o vuota",
            "",
            "Manca il foglio Vendite_2025 oppure è vuoto.",
        ),
        ("File input non trovato", "", "Il file caricato non è stato trovato durante l'esecuzione."),
        ("", "", "Errore durante l'analisi. Controlla il file di input e riprova."),
        ("Core column missing: NetPrice", "", "Core column missing: NetPrice"),
    ],
)
def test_parse_engine_error_covers_known_and_generic_cases(marginrail_module, stderr, stdout, expected):
    assert marginrail_module.parse_engine_error(stderr, stdout) == expected


def test_build_zip_bytes_includes_nested_files(marginrail_module, tmp_path):
    root = tmp_path / "output"
    nested = root / "nested"
    nested.mkdir(parents=True)
    (root / "a.txt").write_text("alpha", encoding="utf-8")
    (nested / "b.txt").write_text("beta", encoding="utf-8")

    zip_bytes = marginrail_module.build_zip_bytes(root)

    with ZipFile(io.BytesIO(zip_bytes)) as archive:
        names = sorted(archive.namelist())
        assert names == ["a.txt", "nested/b.txt"]
        assert archive.read("a.txt") == b"alpha"
        assert archive.read("nested/b.txt") == b"beta"


def test_run_marginrail_success_generates_outputs_and_consistent_results(
    marginrail_module,
    valid_excel_bytes,
    tmp_path,
    monkeypatch,
):
    base_dir = tmp_path / "project"
    (base_dir / "engine").mkdir(parents=True)
    (base_dir / "config").mkdir(parents=True)
    (base_dir / "engine" / "main_engine.py").write_text("# fake engine", encoding="utf-8")
    (base_dir / "config" / "rules_config.json").write_text("{}", encoding="utf-8")

    monkeypatch.setattr(marginrail_module, "BASE_DIR", base_dir)
    monkeypatch.setattr(marginrail_module, "ENGINE_SCRIPT", base_dir / "engine" / "main_engine.py")
    monkeypatch.setattr(marginrail_module, "CONFIG_PATH", base_dir / "config" / "rules_config.json")
    monkeypatch.setattr(marginrail_module, "RUNS_DIR", base_dir / "runs")
    monkeypatch.setattr(marginrail_module, "build_run_id", lambda: "run_test_001")

    def fake_subprocess_run(cmd, capture_output, text):
        assert capture_output is True
        assert text is True
        assert cmd[0] == sys.executable
        input_path = Path(cmd[cmd.index("--input") + 1])
        output_dir = Path(cmd[cmd.index("--output-dir") + 1])
        config_path = Path(cmd[cmd.index("--config") + 1])

        assert input_path.exists()
        assert config_path.exists()

        output_dir.mkdir(parents=True, exist_ok=True)

        kpis = {
            "totale_casi": 2,
            "totale_rischio_eur": 175.0,
            "casi_critical": 1,
            "casi_high": 1,
            "totale_righe_vendita": 10,
            "clienti_coinvolti": 2,
            "ordini_coinvolti": 2,
            "regole_attive": 2,
        }
        (output_dir / "kpi_controllo_margini_v3.json").write_text(
            json.dumps(kpis), encoding="utf-8"
        )

        cases_df = pd.DataFrame(
            [
                {
                    "CaseID": "C-001",
                    "RuleCode": "RULE_A",
                    "MarginRiskEUR": 150.0,
                    "Severity": "Critical",
                    "Owner": "sales",
                    "Cliente": "Acme",
                    "Prodotto": "P1",
                    "NumeroOrdine": "O-1",
                    "Reason": "discount too high",
                    "SuggestedAction": "review",
                },
                {
                    "CaseID": "C-002",
                    "RuleCode": "RULE_B",
                    "MarginRiskEUR": 25.0,
                    "Severity": "High",
                    "Owner": "sales",
                    "Cliente": "Beta",
                    "Prodotto": "P2",
                    "NumeroOrdine": "O-2",
                    "Reason": "price mismatch",
                    "SuggestedAction": "check",
                },
            ]
        )
        cases_df.to_csv(output_dir / "casi_controllo_margini_v3.csv", index=False)

        review_df = pd.DataFrame([
            {"CaseID": "C-001", "ReviewStatus": "OPEN"},
            {"CaseID": "C-002", "ReviewStatus": "OPEN"},
        ])
        review_df.to_csv(output_dir / "review_pack_step1.csv", index=False)

        (output_dir / "report_controllo_margini_v3.xlsx").write_bytes(build_report_xlsx_bytes())
        (output_dir / "config_effettiva_usata.json").write_text(
            json.dumps({"threshold": 10}), encoding="utf-8"
        )

        return subprocess.CompletedProcess(cmd, 0, stdout="engine ok", stderr="")

    monkeypatch.setattr(marginrail_module.subprocess, "run", fake_subprocess_run)

    results = marginrail_module.run_marginrail(valid_excel_bytes, "input.xlsx")

    assert results["run_id"] == "run_test_001"
    assert results["kpis"]["totale_casi"] == len(results["cases_df"]) == 2
    assert results["cases_df"]["MarginRiskEUR"].sum() == pytest.approx(175.0)
    assert results["top_cases_df"].iloc[0]["CaseID"] == "C-001"
    assert results["top_rules_df"].iloc[0]["RuleCode"] == "RULE_A"
    assert set(results["files_bytes"]) == {
        "casi_controllo_margini_v3.csv",
        "review_pack_step1.csv",
        "report_controllo_margini_v3.xlsx",
        "kpi_controllo_margini_v3.json",
        "config_effettiva_usata.json",
    }

    run_dir = base_dir / "runs" / "run_test_001"
    assert run_dir.exists()
    assert (run_dir / "input" / "input.xlsx").exists()
    assert (run_dir / "output" / "casi_controllo_margini_v3.csv").exists()
    assert (run_dir / "run_test_001_output_pack.zip").exists()
    assert not (run_dir / "_tmp").exists()

    metadata = json.loads((run_dir / "metadata.json").read_text(encoding="utf-8"))
    assert metadata["status"] == "success"
    assert metadata["engine_script"] == "engine/main_engine.py"
    assert metadata["config_file"] == "config/rules_config.json"


@pytest.mark.parametrize(
    ("stderr", "expected_message"),
    [
        (
            "Core column missing: NetPrice",
            "Core column missing: NetPrice",
        ),
        (
            "Sheet 'Vendite_2025' non trovata o vuota",
            "Manca il foglio Vendite_2025 oppure è vuoto.",
        ),
    ],
)
def test_run_marginrail_failure_surfaces_engine_errors(
    marginrail_module,
    valid_excel_bytes,
    tmp_path,
    monkeypatch,
    stderr,
    expected_message,
):
    base_dir = tmp_path / "project"
    (base_dir / "engine").mkdir(parents=True)
    (base_dir / "config").mkdir(parents=True)
    (base_dir / "engine" / "main_engine.py").write_text("# fake engine", encoding="utf-8")
    (base_dir / "config" / "rules_config.json").write_text("{}", encoding="utf-8")

    monkeypatch.setattr(marginrail_module, "BASE_DIR", base_dir)
    monkeypatch.setattr(marginrail_module, "ENGINE_SCRIPT", base_dir / "engine" / "main_engine.py")
    monkeypatch.setattr(marginrail_module, "CONFIG_PATH", base_dir / "config" / "rules_config.json")
    monkeypatch.setattr(marginrail_module, "RUNS_DIR", base_dir / "runs")
    monkeypatch.setattr(marginrail_module, "build_run_id", lambda: "run_test_fail")

    def fake_subprocess_run(cmd, capture_output, text):
        return subprocess.CompletedProcess(cmd, 1, stdout="", stderr=stderr)

    monkeypatch.setattr(marginrail_module.subprocess, "run", fake_subprocess_run)

    with pytest.raises(RuntimeError, match=expected_message.replace("(", r"\(").replace(")", r"\)")):
        marginrail_module.run_marginrail(valid_excel_bytes, "input.xlsx")

    run_dir = base_dir / "runs" / "run_test_fail"
    assert run_dir.exists()
    assert json.loads((run_dir / "metadata.json").read_text(encoding="utf-8"))["status"] == "failed"
    assert (run_dir / "stderr.log").read_text(encoding="utf-8") == stderr
    assert not (run_dir / "_tmp").exists()
