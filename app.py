
from __future__ import annotations

import hashlib
import io
import json
import shutil
import subprocess
import sys
import time
import zipfile
from datetime import datetime
from pathlib import Path
from uuid import uuid4

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

CURRENT_FILE_DIR = Path(__file__).resolve().parent
if str(CURRENT_FILE_DIR) not in sys.path:
    sys.path.insert(0, str(CURRENT_FILE_DIR))

from input_validation import WORKBOOK_SCHEMA, format_report_for_user, validate_workbook_bytes

BASE_DIR = Path(__file__).resolve().parent
ENGINE_SCRIPT = BASE_DIR / "engine" / "main_engine.py"
CONFIG_PATH = BASE_DIR / "config" / "rules_config.json"
TEMPLATE_PATH = BASE_DIR / "templates" / "MarginRail_Input_Template_v1.xlsx"
RUNS_DIR = BASE_DIR / "runs"

REQUIRED_SHEETS = set(WORKBOOK_SCHEMA.keys())

DEFAULT_RETENTION_POLICY = {
    "enabled": False,
    "mode": "keep_last_n",
    "keep_last_n": 100,
    "notes": "Retention disattivata di default. Se enabled=true, vengono eliminate le run più vecchie oltre keep_last_n.",
}

ROWS_ANALYZED_KEYS = [
    "righe_analizzate",
    "rows_analyzed",
    "totale_righe_analizzate",
    "totale_righe_vendita",
    "input_rows_analyzed",
]

for _name, _impl in {"divider": lambda *a, **k: None, "json": lambda *a, **k: None, "markdown": lambda *a, **k: None, "warning": lambda *a, **k: None, "tabs": lambda labels, *a, **k: [st for _ in labels], "text_input": lambda *a, **k: "", "slider": lambda *a, **k: k.get("value", (0.0, 0.0))}.items():
    if not hasattr(st, _name):
        setattr(st, _name, _impl)

st.set_page_config(page_title="MarginRail v2", page_icon="📈", layout="wide")


def runs_index_path() -> Path:
    return RUNS_DIR / "_runs_index.json"

def retention_policy_path() -> Path:
    return RUNS_DIR / "_retention_policy.json"

def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def eur(x: float) -> str:
    try:
        return f"€{float(x):,.0f}"
    except Exception:
        return "€0"


def format_seconds(seconds: float | int | None) -> str:
    if seconds is None:
        return "n.d."
    try:
        value = float(seconds)
    except Exception:
        return "n.d."
    if value < 1:
        return f"{value:.2f}s"
    if value < 60:
        return f"{value:.1f}s"
    minutes = int(value // 60)
    rem = value % 60
    return f"{minutes}m {rem:.1f}s"


def format_bytes(num_bytes: int | None) -> str:
    if num_bytes is None:
        return "n.d."
    try:
        value = float(num_bytes)
    except Exception:
        return "n.d."
    units = ["B", "KB", "MB", "GB"]
    idx = 0
    while value >= 1024 and idx < len(units) - 1:
        value /= 1024
        idx += 1
    if idx == 0:
        return f"{int(value)} {units[idx]}"
    return f"{value:.1f} {units[idx]}"


def safe_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def build_zip_bytes(folder: Path) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in folder.rglob("*"):
            if file_path.is_file():
                zf.write(file_path, arcname=file_path.relative_to(folder))
    buffer.seek(0)
    return buffer.getvalue()


def build_run_id() -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"run_{stamp}_{uuid4().hex[:6]}"


def atomic_write_json(path: Path, payload: dict | list) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(path)


def read_json_file(path: Path, default):
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def ensure_runs_support_files() -> None:
    RUNS_DIR.mkdir(parents=True, exist_ok=True)
    if not retention_policy_path().exists():
        atomic_write_json(retention_policy_path(), DEFAULT_RETENTION_POLICY)
    if not runs_index_path().exists():
        atomic_write_json(
            runs_index_path(),
            {
                "schema_version": 1,
                "updated_at": now_iso(),
                "retention_policy_path": str(retention_policy_path().relative_to(BASE_DIR)),
                "total_runs": 0,
                "runs": [],
            },
        )


def validate_excel_structure(uploaded_bytes: bytes) -> dict:
    try:
        wb = load_workbook(filename=io.BytesIO(uploaded_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return {
            "ok": False,
            "message": f"Il file non è un Excel .xlsx valido: {exc}",
            "found": [],
            "missing": sorted(REQUIRED_SHEETS),
        }

    found = wb.sheetnames
    missing = sorted(REQUIRED_SHEETS - set(found))
    if missing:
        return {
            "ok": False,
            "message": "Mancano uno o più fogli obbligatori.",
            "found": found,
            "missing": missing,
        }
    return {
        "ok": True,
        "message": "Struttura fogli valida.",
        "found": found,
        "missing": [],
    }


def detect_sales_rows_from_input(uploaded_bytes: bytes) -> int | None:
    try:
        wb = load_workbook(filename=io.BytesIO(uploaded_bytes), read_only=True, data_only=True)
        if "Vendite_2025" not in wb.sheetnames:
            return None
        ws = wb["Vendite_2025"]
        max_row = ws.max_row or 0
        return max(max_row - 1, 0)
    except Exception:
        return None


def parse_engine_error(stderr: str, stdout: str) -> str:
    text = (stderr or stdout or "").strip()
    if not text:
        return "Errore durante l'analisi. Controlla il file di input e riprova."
    if "Sheet 'Vendite_2025' non trovata o vuota" in text:
        return "Manca il foglio Vendite_2025 oppure è vuoto."
    if "File input non trovato" in text:
        return "Il file caricato non è stato trovato durante l'esecuzione."
    return text


def extract_rows_analyzed(kpis: dict | None, fallback_rows: int | None = None) -> tuple[int | None, str | None]:
    if isinstance(kpis, dict):
        for key in ROWS_ANALYZED_KEYS:
            value = safe_int(kpis.get(key))
            if value is not None:
                return value, f"kpi:{key}"
    if fallback_rows is not None:
        return fallback_rows, "input_sheet:Vendite_2025"
    return None, None


def build_output_manifest(output_dir: Path) -> list[dict]:
    manifest = []
    if not output_dir.exists():
        return manifest
    for file_path in sorted(output_dir.rglob("*")):
        if file_path.is_file():
            manifest.append(
                {
                    "relative_path": str(file_path.relative_to(output_dir)),
                    "size_bytes": file_path.stat().st_size,
                }
            )
    return manifest


def load_kpis_from_output(output_dir: Path) -> dict:
    kpi_path = output_dir / "kpi_controllo_margini_v3.json"
    if not kpi_path.exists():
        return {}
    try:
        with kpi_path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def load_cases_count_from_output(output_dir: Path) -> int | None:
    cases_path = output_dir / "casi_controllo_margini_v3.csv"
    if not cases_path.exists():
        return None
    try:
        cases_df = pd.read_csv(cases_path)
        return int(len(cases_df))
    except Exception:
        return None


def copy_output_tree(source_dir: Path, target_dir: Path) -> None:
    if not source_dir.exists():
        return
    for child in source_dir.iterdir():
        target = target_dir / child.name
        if child.is_file():
            shutil.copy2(child, target)
        elif child.is_dir():
            if target.exists():
                shutil.rmtree(target)
            shutil.copytree(child, target)


def build_run_metadata(
    run_id: str,
    original_name: str,
    input_bytes: bytes,
    status: str,
    started_at: str,
    finished_at: str,
    duration_seconds: float,
    proc: subprocess.CompletedProcess[str],
    output_dir: Path,
    input_sales_rows_detected: int | None,
) -> dict:
    kpis = load_kpis_from_output(output_dir)
    rows_analyzed, rows_analyzed_source = extract_rows_analyzed(kpis, fallback_rows=input_sales_rows_detected)
    cases_count = load_cases_count_from_output(output_dir)
    output_manifest = build_output_manifest(output_dir)

    metadata = {
        "run_id": run_id,
        "engine_script": str(ENGINE_SCRIPT.relative_to(BASE_DIR)),
        "config_file": str(CONFIG_PATH.relative_to(BASE_DIR)),
        "status": status,
        "started_at": started_at,
        "finished_at": finished_at,
        "duration_seconds": round(float(duration_seconds), 3),
        "duration_human": format_seconds(duration_seconds),
        "input": {
            "file_name": original_name,
            "size_bytes": len(input_bytes),
            "size_human": format_bytes(len(input_bytes)),
            "sha256": sha256_bytes(input_bytes),
            "sales_rows_detected": input_sales_rows_detected,
        },
        "analysis": {
            "rows_analyzed": rows_analyzed,
            "rows_analyzed_source": rows_analyzed_source,
            "cases_count": cases_count,
            "kpi_summary": {
                "totale_casi": safe_int(kpis.get("totale_casi")),
                "totale_rischio_eur": kpis.get("totale_rischio_eur"),
                "casi_critical": safe_int(kpis.get("casi_critical")),
                "casi_high": safe_int(kpis.get("casi_high")),
                "clienti_coinvolti": safe_int(kpis.get("clienti_coinvolti")),
                "ordini_coinvolti": safe_int(kpis.get("ordini_coinvolti")),
                "regole_attive": safe_int(kpis.get("regole_attive")),
            },
        },
        "artifacts": {
            "output_files_count": len(output_manifest),
            "output_manifest": output_manifest,
            "zip_file_name": f"{run_id}_output_pack.zip",
        },
        "execution": {
            "return_code": proc.returncode,
            "engine_script": str(ENGINE_SCRIPT.relative_to(BASE_DIR)),
            "config_file": str(CONFIG_PATH.relative_to(BASE_DIR)),
            "command": [
                sys.executable,
                str(ENGINE_SCRIPT.relative_to(BASE_DIR)),
                "--input",
                f"runs/{run_id}/_tmp/{original_name}",
                "--output-dir",
                f"runs/{run_id}/_tmp/output",
                "--config",
                str(CONFIG_PATH.relative_to(BASE_DIR)),
            ],
            "stdout_log": "stdout.log",
            "stderr_log": "stderr.log",
        },
        "retention": read_json_file(retention_policy_path(), DEFAULT_RETENTION_POLICY),
    }
    return metadata


def build_index_entry(run_dir: Path, metadata: dict) -> dict:
    input_meta = metadata.get("input", {})
    analysis_meta = metadata.get("analysis", {})
    artifacts_meta = metadata.get("artifacts", {})
    return {
        "run_id": metadata.get("run_id", run_dir.name),
        "status": metadata.get("status"),
        "started_at": metadata.get("started_at"),
        "finished_at": metadata.get("finished_at"),
        "duration_seconds": metadata.get("duration_seconds"),
        "input_file": input_meta.get("file_name"),
        "input_size_bytes": input_meta.get("size_bytes"),
        "input_sha256": input_meta.get("sha256"),
        "rows_analyzed": analysis_meta.get("rows_analyzed"),
        "cases_count": analysis_meta.get("cases_count"),
        "output_files_count": artifacts_meta.get("output_files_count"),
        "run_path": str(run_dir.relative_to(BASE_DIR)),
    }


def read_run_metadata(run_dir: Path) -> dict | None:
    metadata_path = run_dir / "metadata.json"
    if not metadata_path.exists():
        return None
    try:
        return json.loads(metadata_path.read_text(encoding="utf-8"))
    except Exception:
        return None


def rebuild_runs_index_from_disk() -> dict:
    ensure_runs_support_files()
    entries = []
    for run_dir in sorted(RUNS_DIR.glob("run_*")):
        if not run_dir.is_dir():
            continue
        metadata = read_run_metadata(run_dir)
        if not metadata:
            continue
        entries.append(build_index_entry(run_dir, metadata))

    entries = sorted(
        entries,
        key=lambda x: ((x.get("started_at") or ""), (x.get("run_id") or "")),
        reverse=True,
    )
    index_payload = {
        "schema_version": 1,
        "updated_at": now_iso(),
        "retention_policy_path": str(retention_policy_path().relative_to(BASE_DIR)),
        "total_runs": len(entries),
        "runs": entries,
    }
    atomic_write_json(runs_index_path(), index_payload)
    return index_payload


def upsert_run_index(run_dir: Path, metadata: dict) -> dict:
    ensure_runs_support_files()
    index_payload = read_json_file(runs_index_path(), None)
    if not isinstance(index_payload, dict) or "runs" not in index_payload:
        index_payload = rebuild_runs_index_from_disk()

    entry = build_index_entry(run_dir, metadata)
    existing = [x for x in index_payload.get("runs", []) if x.get("run_id") != entry["run_id"]]
    existing.append(entry)
    existing = sorted(
        existing,
        key=lambda x: ((x.get("started_at") or ""), (x.get("run_id") or "")),
        reverse=True,
    )
    index_payload = {
        "schema_version": 1,
        "updated_at": now_iso(),
        "retention_policy_path": str(retention_policy_path().relative_to(BASE_DIR)),
        "total_runs": len(existing),
        "runs": existing,
    }
    atomic_write_json(runs_index_path(), index_payload)
    return index_payload


def apply_retention_policy_if_enabled() -> dict:
    ensure_runs_support_files()
    policy = read_json_file(retention_policy_path(), DEFAULT_RETENTION_POLICY)
    index_payload = read_json_file(runs_index_path(), None)
    if not isinstance(index_payload, dict) or "runs" not in index_payload:
        index_payload = rebuild_runs_index_from_disk()

    result = {
        "enabled": bool(policy.get("enabled", False)),
        "mode": policy.get("mode", "keep_last_n"),
        "deleted_runs": [],
        "kept_runs": index_payload.get("total_runs", 0),
    }

    if not result["enabled"]:
        return result

    if result["mode"] != "keep_last_n":
        result["error"] = f"Modalità retention non supportata: {result['mode']}"
        return result

    keep_last_n = max(int(policy.get("keep_last_n", 100)), 1)
    runs = index_payload.get("runs", [])
    to_delete = runs[keep_last_n:]

    deleted_runs = []
    for entry in to_delete:
        run_path = entry.get("run_path")
        if not run_path:
            continue
        abs_path = BASE_DIR / run_path
        if abs_path.exists() and abs_path.is_dir():
            shutil.rmtree(abs_path, ignore_errors=True)
            deleted_runs.append(entry.get("run_id"))

    if deleted_runs:
        index_payload = rebuild_runs_index_from_disk()

    result["deleted_runs"] = deleted_runs
    result["kept_runs"] = index_payload.get("total_runs", 0)
    return result


def persist_run_files(
    run_dir: Path,
    input_bytes: bytes,
    original_name: str,
    proc: subprocess.CompletedProcess[str],
    output_dir: Path,
    metadata: dict,
) -> dict:
    run_dir.mkdir(parents=True, exist_ok=True)
    (run_dir / "input").mkdir(exist_ok=True)
    (run_dir / "output").mkdir(exist_ok=True)

    input_path = run_dir / "input" / original_name
    input_path.write_bytes(input_bytes)

    copy_output_tree(output_dir, run_dir / "output")

    atomic_write_json(run_dir / "metadata.json", metadata)
    (run_dir / "stdout.log").write_text(proc.stdout or "", encoding="utf-8")
    (run_dir / "stderr.log").write_text(proc.stderr or "", encoding="utf-8")

    return upsert_run_index(run_dir, metadata)


def run_marginrail(uploaded_bytes: bytes, original_name: str) -> dict:
    ensure_runs_support_files()

    structure_check = validate_excel_structure(uploaded_bytes)
    if not structure_check["ok"]:
        if "Vendite_2025" in structure_check.get("missing", []):
            raise RuntimeError(parse_engine_error("Sheet 'Vendite_2025' non trovata o vuota", ""))
        raise RuntimeError(structure_check["message"])

    run_id = build_run_id()
    run_dir = RUNS_DIR / run_id
    tmp_dir = run_dir / "_tmp"
    output_dir = tmp_dir / "output"
    tmp_dir.mkdir(parents=True, exist_ok=True)

    started_at = now_iso()
    started_perf = time.perf_counter()
    input_sales_rows_detected = detect_sales_rows_from_input(uploaded_bytes)

    input_path = tmp_dir / original_name
    input_path.write_bytes(uploaded_bytes)

    cmd = [
        sys.executable,
        str(ENGINE_SCRIPT),
        "--input",
        str(input_path),
        "--output-dir",
        str(output_dir),
        "--config",
        str(CONFIG_PATH),
    ]

    proc = subprocess.run(cmd, capture_output=True, text=True)
    finished_at = now_iso()
    duration_seconds = time.perf_counter() - started_perf
    required_outputs_ok = (
        (output_dir / "kpi_controllo_margini_v3.json").exists()
        and (output_dir / "casi_controllo_margini_v3.csv").exists()
    )
    status = "success" if proc.returncode == 0 and required_outputs_ok else "failed"

    metadata = build_run_metadata(
        run_id=run_id,
        original_name=original_name,
        input_bytes=uploaded_bytes,
        status=status,
        started_at=started_at,
        finished_at=finished_at,
        duration_seconds=duration_seconds,
        proc=proc,
        output_dir=output_dir,
        input_sales_rows_detected=input_sales_rows_detected,
    )

    persist_run_files(
        run_dir=run_dir,
        input_bytes=uploaded_bytes,
        original_name=original_name,
        proc=proc,
        output_dir=output_dir,
        metadata=metadata,
    )

    retention_result = apply_retention_policy_if_enabled()

    if tmp_dir.exists():
        shutil.rmtree(tmp_dir, ignore_errors=True)

    if proc.returncode != 0:
        raise RuntimeError(parse_engine_error(proc.stderr, proc.stdout))
    if not required_outputs_ok:
        raise RuntimeError("Output non generati correttamente.")

    output_saved_dir = run_dir / "output"
    kpi_path = output_saved_dir / "kpi_controllo_margini_v3.json"
    cases_path = output_saved_dir / "casi_controllo_margini_v3.csv"
    review_path = output_saved_dir / "review_pack_step1.csv"
    report_path = output_saved_dir / "report_controllo_margini_v3.xlsx"
    config_used_path = output_saved_dir / "config_effettiva_usata.json"

    with kpi_path.open("r", encoding="utf-8") as f:
        kpis = json.load(f)

    cases_df = pd.read_csv(cases_path)
    review_df = pd.read_csv(review_path) if review_path.exists() else pd.DataFrame()

    top_rules = (
        cases_df.groupby("RuleCode", dropna=False)
        .agg(Casi=("CaseID", "count"), RischioEUR=("MarginRiskEUR", "sum"))
        .reset_index()
        .sort_values(["RischioEUR", "Casi"], ascending=[False, False])
    )

    top_cases = cases_df.sort_values("MarginRiskEUR", ascending=False).head(50)

    zip_bytes = build_zip_bytes(output_saved_dir)
    zip_path = run_dir / f"{run_id}_output_pack.zip"
    zip_path.write_bytes(zip_bytes)

    files_bytes = {}
    for p in [cases_path, review_path, report_path, kpi_path, config_used_path]:
        if p.exists():
            files_bytes[p.name] = p.read_bytes()

    metadata = read_run_metadata(run_dir) or metadata
    index_payload = read_json_file(runs_index_path(), None)
    if not isinstance(index_payload, dict) or "runs" not in index_payload:
        index_payload = rebuild_runs_index_from_disk()

    return {
        "run_id": run_id,
        "run_dir": str(run_dir),
        "run_dir_relative": str(run_dir.relative_to(BASE_DIR)),
        "metadata": metadata,
        "kpis": kpis,
        "cases_df": cases_df,
        "review_df": review_df,
        "top_rules_df": top_rules,
        "top_cases_df": top_cases,
        "zip_bytes": zip_bytes,
        "zip_path": str(zip_path),
        "files_bytes": files_bytes,
        "stdout": proc.stdout,
        "stderr": proc.stderr,
        "retention_result": retention_result,
        "runs_index": index_payload,
    }


def load_runs_index() -> dict:
    ensure_runs_support_files()
    index_payload = read_json_file(runs_index_path(), None)
    if not isinstance(index_payload, dict) or "runs" not in index_payload:
        index_payload = rebuild_runs_index_from_disk()
    return index_payload


def load_retention_policy() -> dict:
    ensure_runs_support_files()
    policy = read_json_file(retention_policy_path(), DEFAULT_RETENTION_POLICY)
    if not isinstance(policy, dict):
        policy = DEFAULT_RETENTION_POLICY.copy()
    return policy


def build_runs_history_dataframe(index_payload: dict) -> pd.DataFrame:
    rows = []
    for entry in index_payload.get("runs", []):
        rows.append(
            {
                "Run ID": entry.get("run_id"),
                "Status": entry.get("status"),
                "Started": entry.get("started_at"),
                "Duration": format_seconds(entry.get("duration_seconds")),
                "Input file": entry.get("input_file"),
                "Input size": format_bytes(entry.get("input_size_bytes")),
                "Rows analyzed": entry.get("rows_analyzed"),
                "Cases": entry.get("cases_count"),
                "Output files": entry.get("output_files_count"),
                "SHA256 (short)": (entry.get("input_sha256") or "")[:12],
                "Path": entry.get("run_path"),
            }
        )
    return pd.DataFrame(rows)


def render_run_metadata(metadata: dict) -> None:
    input_meta = metadata.get("input", {})
    analysis_meta = metadata.get("analysis", {})
    execution_meta = metadata.get("execution", {})

    st.subheader("Dettagli run salvata")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Durata run", metadata.get("duration_human", "n.d."))
    m2.metric("Dimensione input", input_meta.get("size_human", "n.d."))
    m3.metric("Righe analizzate", f"{safe_int(analysis_meta.get('rows_analyzed')):,}" if safe_int(analysis_meta.get("rows_analyzed")) is not None else "n.d.")
    m4.metric("Return code", execution_meta.get("return_code", "n.d."))

    st.caption(
        f"Input: {input_meta.get('file_name', 'n.d.')} • "
        f"SHA256: {(input_meta.get('sha256') or '')[:16]}… • "
        f"Inizio: {metadata.get('started_at', 'n.d.')} • "
        f"Fine: {metadata.get('finished_at', 'n.d.')}"
    )

    with st.expander("metadata.json completo"):
        st.json(metadata)


def render_runs_history() -> None:
    index_payload = load_runs_index()
    policy = load_retention_policy()

    st.subheader("Storico run locali")
    runs_df = build_runs_history_dataframe(index_payload)

    h1, h2, h3 = st.columns(3)
    h1.metric("Run totali", int(index_payload.get("total_runs", 0)))
    h2.metric(
        "Retention",
        "attiva" if policy.get("enabled") else "disattivata",
        help=f"Modalità: {policy.get('mode')} • keep_last_n: {policy.get('keep_last_n')}",
    )
    status_counts = runs_df["Status"].value_counts().to_dict() if not runs_df.empty else {}
    h3.metric("Run riuscite", int(status_counts.get("success", 0)))

    if runs_df.empty:
        st.info("Nessuna run salvata in runs/.")
        return

    st.dataframe(runs_df.head(15), use_container_width=True, hide_index=True, height=360)

    with st.expander("Indice locale run (_runs_index.json)"):
        st.json(index_payload)

    with st.expander("Policy retention (_retention_policy.json)"):
        st.json(policy)


def render_results(results: dict) -> None:
    kpis = results["kpis"]
    cases_df = results["cases_df"]
    top_rules_df = results["top_rules_df"]
    metadata = results.get("metadata", {})

    st.success("Analisi completata.")
    st.caption(f"Run ID: {results['run_id']} • Salvata in: {results['run_dir_relative']}")

    render_run_metadata(metadata)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Casi totali", f"{int(kpis.get('totale_casi', 0)):,}")
    c2.metric("Rischio totale", eur(kpis.get("totale_rischio_eur", 0)))
    c3.metric("Critical", f"{int(kpis.get('casi_critical', 0)):,}")
    c4.metric("High", f"{int(kpis.get('casi_high', 0)):,}")

    c5, c6, c7, c8 = st.columns(4)
    c5.metric("Righe vendita", f"{int(kpis.get('totale_righe_vendita', 0)):,}")
    c6.metric("Clienti coinvolti", f"{int(kpis.get('clienti_coinvolti', 0)):,}")
    c7.metric("Ordini coinvolti", f"{int(kpis.get('ordini_coinvolti', 0)):,}")
    c8.metric("Regole attive", f"{int(kpis.get('regole_attive', 0)):,}")

    st.subheader("Top regole per rischio")
    st.dataframe(top_rules_df.head(10), use_container_width=True, hide_index=True)

    st.subheader("Top casi")
    severity_values = sorted([x for x in cases_df["Severity"].dropna().unique().tolist()])
    rule_values = sorted([x for x in cases_df["RuleCode"].dropna().unique().tolist()])

    f1, f2 = st.columns(2)
    selected_severity = f1.multiselect("Severity", severity_values, default=severity_values)
    selected_rules = f2.multiselect("RuleCode", rule_values, default=rule_values[: min(6, len(rule_values))])

    filtered = cases_df.copy()
    if selected_severity:
        filtered = filtered[filtered["Severity"].isin(selected_severity)]
    if selected_rules:
        filtered = filtered[filtered["RuleCode"].isin(selected_rules)]

    filtered = filtered.sort_values("MarginRiskEUR", ascending=False)
    show_cols = [
        c for c in [
            "CaseID", "RuleCode", "Severity", "Owner", "MarginRiskEUR",
            "Cliente", "Prodotto", "NumeroOrdine", "Reason", "SuggestedAction"
        ] if c in filtered.columns
    ]
    st.dataframe(filtered[show_cols], use_container_width=True, hide_index=True, height=450)

    st.subheader("Download output")
    col_a, col_b, col_c = st.columns(3)
    col_a.download_button(
        "Scarica pacchetto completo (.zip)",
        data=results["zip_bytes"],
        file_name="marginrail_output_pack.zip",
        mime="application/zip",
        use_container_width=True,
    )
    files_bytes = results["files_bytes"]
    if "report_controllo_margini_v3.xlsx" in files_bytes:
        col_b.download_button(
            "Scarica report Excel",
            data=files_bytes["report_controllo_margini_v3.xlsx"],
            file_name="MarginRail_Executive_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    if "casi_controllo_margini_v3.csv" in files_bytes:
        col_c.download_button(
            "Scarica casi CSV",
            data=files_bytes["casi_controllo_margini_v3.csv"],
            file_name="MarginRail_Cases.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with st.expander("Log tecnico run"):
        st.code(results["stdout"] or "Nessun log disponibile.")

    if results.get("retention_result", {}).get("enabled"):
        deleted = results["retention_result"].get("deleted_runs", [])
        if deleted:
            st.warning(f"Retention applicata: eliminate {len(deleted)} run vecchie.")
        else:
            st.info("Retention attiva, nessuna run eliminata in questa esecuzione.")


ensure_runs_support_files()

def _can_render_full_ui() -> bool:
    try:
        return hasattr(st.columns(1)[0], "metric")
    except Exception:
        return False

if _can_render_full_ui():
    st.title("MarginRail v2")
    st.caption("Upload Excel standardizzato → analisi automatica → dashboard + output persistenti")

    left, right = st.columns([2, 1])

    with left:
        st.subheader("1) Carica il file Excel")
        uploaded_file = st.file_uploader("Excel cliente", type=["xlsx"])
        precheck = None
        if uploaded_file is not None:
            uploaded_bytes = uploaded_file.getvalue()
            precheck = validate_excel_structure(uploaded_bytes)
            if precheck["ok"]:
                st.success("Controllo struttura: OK")
                detected_rows = detect_sales_rows_from_input(uploaded_bytes)
                if detected_rows is not None:
                    st.caption(f"Righe rilevate in Vendite_2025: {detected_rows:,}")
            else:
                st.error(precheck["message"])
                if precheck["missing"]:
                    st.write("Fogli mancanti:", ", ".join(precheck["missing"]))
        run_clicked = st.button(
            "Esegui analisi",
            type="primary",
            use_container_width=True,
            disabled=(uploaded_file is None or (precheck is not None and not precheck["ok"]))
        )

    with right:
        st.subheader("Template input")
        if TEMPLATE_PATH.exists():
            st.download_button(
                "Scarica template Excel",
                data=TEMPLATE_PATH.read_bytes(),
                file_name=TEMPLATE_PATH.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        st.info(
            "Il file deve usare il template ufficiale con i fogli richiesti. "
            "La V2 salva ogni run su filesystem locale in runs/, con metadata.json, log, indice e policy retention."
        )

    if run_clicked and uploaded_file is not None:
        with st.spinner("Sto eseguendo l'analisi..."):
            try:
                st.session_state["mr_results"] = run_marginrail(uploaded_file.getvalue(), uploaded_file.name)
            except Exception as e:
                st.error(str(e))

    if "mr_results" in st.session_state:
        render_results(st.session_state["mr_results"])

    st.divider()
    render_runs_history()
