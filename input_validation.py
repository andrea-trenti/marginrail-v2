from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


@dataclass(frozen=True)
class SheetSchema:
    core_columns: tuple[str, ...]
    optional_columns: tuple[str, ...] = ()
    require_data_rows: bool = False


@dataclass
class ValidationReport:
    ok: bool
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    found_sheets: List[str] = field(default_factory=list)
    missing_sheets: List[str] = field(default_factory=list)
    missing_core_by_sheet: Dict[str, List[str]] = field(default_factory=dict)
    missing_optional_by_sheet: Dict[str, List[str]] = field(default_factory=dict)
    empty_optional_sheets: List[str] = field(default_factory=list)
    duplicated_headers_by_sheet: Dict[str, List[str]] = field(default_factory=dict)

    @property
    def blocking_message(self) -> str:
        if self.ok:
            return "File leggibile e struttura core valida."
        return "Il file non può essere analizzato finché i problemi bloccanti non vengono corretti."

    @property
    def warning_message(self) -> str:
        if not self.warnings:
            return ""
        return "Sono presenti avvisi non bloccanti: l’analisi può partire, ma conviene correggerli per avere output più completi."

    def to_dict(self) -> dict:
        return {
            "ok": self.ok,
            "errors": self.errors,
            "warnings": self.warnings,
            "blocking_message": self.blocking_message,
            "warning_message": self.warning_message,
            "found_sheets": self.found_sheets,
            "missing_sheets": self.missing_sheets,
            "missing_core_by_sheet": self.missing_core_by_sheet,
            "missing_optional_by_sheet": self.missing_optional_by_sheet,
            "empty_optional_sheets": self.empty_optional_sheets,
            "duplicated_headers_by_sheet": self.duplicated_headers_by_sheet,
        }


class InputValidationError(RuntimeError):
    """Errore business-friendly per input non valido."""


WORKBOOK_SCHEMA: Dict[str, SheetSchema] = {
    "Vendite_2025": SheetSchema(
        core_columns=(
            "NumeroOrdine",
            "RigaOrdine",
            "TipoDocumento",
            "ClienteID",
            "CanaleVendita",
            "ProdottoID",
            "QtaDocumento",
            "QtaOrdinata",
            "PrezzoListinoUnit",
            "ScontoBasePct",
            "ScontoContrattualePct",
            "ScontoPromoPct",
            "ScontoExtraPct",
            "ScontoTotalePct",
            "PrezzoNettoUnit",
            "FloorPriceUnit",
            "CostoAttualeUnit",
            "VariazioneCostoPct",
            "RicavoRiga",
            "MargineContributivo",
            "MarginePct",
            "PromoID",
            "AccordoID",
            "MotivoDeroga",
            "DataDocumento",
        ),
        optional_columns=(
            "Cliente",
            "Venditore",
            "Prodotto",
            "Categoria",
            "QtaResa",
            "DataOrdine",
            "PrioritaOrdine",
            "GruppoCliente",
            "Regione",
            "Provincia",
        ),
        require_data_rows=True,
    ),
    "Workflow_Deroghe_2025": SheetSchema(
        core_columns=(
            "NumeroOrdine",
            "RigaOrdine",
            "SogliaScontoRuoloPct",
            "StatoDeroga",
        ),
        optional_columns=(
            "ApprovatoDa",
            "DataApprovazione",
            "MotivoDeroga",
        ),
    ),
    "Clienti": SheetSchema(
        core_columns=(
            "ClienteID",
            "RischioCredito",
        ),
        optional_columns=(
            "ScontoBase",
            "AreaCommerciale",
        ),
    ),
    "Prodotti": SheetSchema(
        core_columns=(
            "ProdottoID",
            "MargineTarget",
        ),
        optional_columns=(
            "ClasseBrand",
            "StatoProdotto",
        ),
    ),
    "Accordi_Commerciali": SheetSchema(
        core_columns=(
            "AccordoID",
            "FloorPriceUnit",
            "StatoAccordo",
            "ValidoDa",
            "ValidoA",
        ),
        optional_columns=(
            "PrezzoContrattualeUnit",
            "ScontoContrattualePct",
        ),
    ),
    "Promo_2025": SheetSchema(
        core_columns=(
            "PromoID",
            "DataInizio",
            "DataFine",
            "CanaleValido",
        ),
        optional_columns=(
            "ScontoExtraPct",
            "MotivoPromo",
        ),
    ),
}


def ensure_optional_columns(df, columns: Iterable[str]):
    for col in columns:
        if col not in df.columns:
            df[col] = None
    return df


def _clean_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _extract_headers(ws) -> List[str]:
    row_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    first_row = next(row_iter, None)
    if first_row is None:
        return []
    return [_clean_header(cell) for cell in first_row if _clean_header(cell)]


def _sheet_has_data_rows(ws) -> bool:
    row_iter = ws.iter_rows(min_row=2, values_only=True)
    for row in row_iter:
        if row is None:
            continue
        if any(cell not in (None, "") for cell in row):
            return True
    return False


def _find_duplicates(values: Sequence[str]) -> List[str]:
    seen = set()
    duplicates = []
    for value in values:
        if value in seen and value not in duplicates:
            duplicates.append(value)
        seen.add(value)
    return duplicates


def _build_report_from_workbook(wb) -> ValidationReport:
    report = ValidationReport(ok=True)
    report.found_sheets = list(wb.sheetnames)

    found_set = set(wb.sheetnames)
    required_set = set(WORKBOOK_SCHEMA.keys())
    report.missing_sheets = sorted(required_set - found_set)

    if report.missing_sheets:
        report.ok = False
        report.errors.append(
            "Mancano uno o più fogli obbligatori: "
            + ", ".join(report.missing_sheets)
            + "."
        )

    for sheet_name, schema in WORKBOOK_SCHEMA.items():
        if sheet_name not in found_set:
            continue

        ws = wb[sheet_name]
        headers = _extract_headers(ws)
        duplicates = _find_duplicates(headers)
        if duplicates:
            report.ok = False
            report.duplicated_headers_by_sheet[sheet_name] = duplicates
            report.errors.append(
                f"Foglio {sheet_name}: intestazioni duplicate trovate ({', '.join(duplicates)})."
            )

        if not headers:
            report.ok = False
            report.missing_core_by_sheet[sheet_name] = list(schema.core_columns)
            report.errors.append(
                f"Foglio {sheet_name}: intestazioni mancanti o foglio vuoto."
            )
            continue

        missing_core = [col for col in schema.core_columns if col not in headers]
        missing_optional = [col for col in schema.optional_columns if col not in headers]

        if missing_core:
            report.ok = False
            report.missing_core_by_sheet[sheet_name] = missing_core
            report.errors.append(
                f"Foglio {sheet_name}: mancano colonne core ({', '.join(missing_core)})."
            )

        if missing_optional:
            report.missing_optional_by_sheet[sheet_name] = missing_optional
            report.warnings.append(
                f"Foglio {sheet_name}: mancano colonne opzionali ({', '.join(missing_optional)})."
            )

        has_data_rows = _sheet_has_data_rows(ws)
        if schema.require_data_rows and not has_data_rows:
            report.ok = False
            report.errors.append(
                f"Foglio {sheet_name}: presente ma senza righe dati. Serve almeno una riga per avviare l’analisi."
            )
        elif not schema.require_data_rows and not has_data_rows:
            report.empty_optional_sheets.append(sheet_name)
            report.warnings.append(
                f"Foglio {sheet_name}: presente ma senza righe dati. L’analisi parte comunque, ma alcune regole potrebbero non attivarsi."
            )

    return report


def validate_workbook_bytes(uploaded_bytes: bytes) -> ValidationReport:
    try:
        wb = load_workbook(
            filename=io.BytesIO(uploaded_bytes),
            read_only=True,
            data_only=True,
        )
    except (InvalidFileException, zipfile.BadZipFile):
        return ValidationReport(
            ok=False,
            errors=[
                "Il file caricato non è un vero .xlsx leggibile. Scarica di nuovo il template ufficiale e riprova."
            ],
        )
    except Exception as exc:
        return ValidationReport(
            ok=False,
            errors=[
                "Il file .xlsx non è leggibile oppure è danneggiato.",
                f"Dettaglio sintetico: {exc}",
            ],
        )

    try:
        return _build_report_from_workbook(wb)
    finally:
        wb.close()


def validate_workbook_file(file_path: Path) -> ValidationReport:
    try:
        wb = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
        )
    except (InvalidFileException, zipfile.BadZipFile):
        return ValidationReport(
            ok=False,
            errors=[
                "Il file selezionato non è un vero .xlsx leggibile. Usa il template ufficiale MarginRail."
            ],
        )
    except Exception as exc:
        return ValidationReport(
            ok=False,
            errors=[
                "Il file .xlsx non è leggibile oppure è danneggiato.",
                f"Dettaglio sintetico: {exc}",
            ],
        )

    try:
        return _build_report_from_workbook(wb)
    finally:
        wb.close()


def format_report_for_user(report: ValidationReport) -> str:
    lines: List[str] = []

    if report.errors:
        lines.append("Problemi bloccanti:")
        lines.extend(f"- {message}" for message in report.errors)

    if report.warnings:
        if lines:
            lines.append("")
        lines.append("Avvisi non bloccanti:")
        lines.extend(f"- {message}" for message in report.warnings)

    if not lines:
        lines.append("File leggibile e struttura input valida.")

    return "\n".join(lines)


def raise_if_invalid_file(file_path: Path) -> ValidationReport:
    report = validate_workbook_file(file_path)
    if not report.ok:
        raise InputValidationError(format_report_for_user(report))
    return report
